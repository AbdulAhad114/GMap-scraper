from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import time
import re
import csv
import openpyxl


# def scrape_google_maps_data(search_query, output_filename):
#     options = webdriver.ChromeOptions()
#     options.add_argument('--no-sandbox')
#     options.add_argument('--disable-dev-shm-usage')
#     options.add_argument('--headless')
#     driver = webdriver.Chrome(options=options)

#     # Navigate to Google Maps
#     driver.get(f'https://www.google.com/maps/search/{search_query}')

#     results_container_selector = '.m6QErb.DxyBCb.kA9KIf.dS8AEf.XiKgde.ecceSd'
#     WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, results_container_selector)))

#     business_urls = set()

#     scroll_and_collect(driver, business_urls, results_container_selector)

#     data = []

#     for business_url in business_urls:
#         business_data = scrape_business_data(driver, business_url)
#         data.append(business_data)
#         print('Scraped Business Data:', business_data)

#     # Save data to CSV
#     save_to_csv(data, output_filename)

#     driver.quit()

def scrape_google_maps_data(search_query, output_filename):
    options = webdriver.ChromeOptions()
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--headless')
    driver = webdriver.Chrome()

    # Navigate to Google Maps
    driver.get(f'https://www.google.com/maps/search/{search_query}')

    # CSS selector for the container that lists search results
    results_container_selector = '.m6QErb.DxyBCb.kA9KIf.dS8AEf.XiKgde.ecceSd'

    # Wait for search results to load
    try:
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, results_container_selector)))
        business_urls = set()

        # Call a function to scroll and collect URLs
        scroll_and_collect(driver, business_urls, results_container_selector)

        # Check if no URLs were collected and print a message
        if not business_urls:
            print(f"No results found for: {search_query}. Moving to the next business type.")
            driver.quit()
            return

        data = []
        for business_url in business_urls:
            business_data = scrape_business_data(driver, business_url)
            data.append(business_data)
            print('Scraped Business Data:', business_data)

        # Save data to CSV
        save_to_csv(data, output_filename)

    except Exception as e:
        print(f"Could not collect data for {search_query}: {e}")
    
    finally:
        driver.quit()


def scroll_and_collect(driver, business_urls, results_container_selector):
    try:
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, results_container_selector))
        )

        results_container = driver.find_element(By.CSS_SELECTOR, results_container_selector)
        new_urls = results_container.find_elements(By.CSS_SELECTOR, 'a[href^="https://www.google.com/maps/place"]')
        for anchor in new_urls:
            business_urls.add(anchor.get_attribute('href'))

        while True:
            driver.execute_script("return arguments[0].scrollIntoView();", new_urls[-1])
            time.sleep(2)

            results_container = driver.find_element(By.CSS_SELECTOR, results_container_selector)
            new_urls = results_container.find_elements(By.CSS_SELECTOR, 'a[href^="https://www.google.com/maps/place"]')

            for anchor in new_urls:
                business_urls.add(anchor.get_attribute('href'))

            print(f'Collected {len(new_urls)} new URLs, total: {len(business_urls)}')
            page_text = driver.find_element(By.TAG_NAME, 'body').text
            end_list_string = "You've reached the end of the list."
            if end_list_string in page_text:
                break

            if not new_urls:
                break

    except Exception as e:
        print(f"An error occurred: {e}")

    return business_urls


def scrape_business_data(driver, business_url):
    try:
        print(f'Fetching data from: {business_url}')
        driver.get(business_url)
        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'h1.DUwDvf.lfPIob')))

        soup = BeautifulSoup(driver.page_source, 'html.parser')
        
        h1_element = soup.find('h1', class_='DUwDvf lfPIob')
        name = h1_element.get_text(strip=True) if h1_element else 'N/A'

        reviews_element = soup.find('span', attrs={'aria-label': re.compile(r'reviews')})
        number_of_reviews = reviews_element.get_text(strip=True).split('(')[1].replace(')', '').replace(',', '') if reviews_element else 'N/A'

        ratings_element = soup.find('span', class_='ceNzKf', role='img')
        ratings = ratings_element.get('aria-label').split(' ')[0] if ratings_element else 'N/A'

        address = soup.find('div', class_='Io6YTe fontBodyMedium kR99db fdkmkc').get_text(strip=True) if soup.find('div', class_='Io6YTe fontBodyMedium kR99db fdkmkc') else 'N/A'

        phone_divs = soup.find_all('div', class_='Io6YTe fontBodyMedium kR99db fdkmkc')
        phone = 'N/A'
        phone_regex = r'(\+?\d{1,2}\s*[-.\s]?)?(\(?\d{3}\)?[\s.-]?)?\d{3}[\s.-]?\d{4}'

        for div in phone_divs:
            text = div.get_text(strip=True)
            if re.match(phone_regex, text):
                phone = text
                break

        website_elements = soup.find_all('div', class_='rogA2c ITvuef')
        website = next((el.get_text(strip=True) for div in website_elements for el in div.find_all('div', class_='Io6YTe') if '.' in el.get_text()), 'N/A')

        return {
            'name': name,
            'ratings': ratings,
            'numberOfReviews': number_of_reviews,
            'phone': phone,
            'website': website,
            'address': address,
            'link': business_url
        }
    except Exception as e:
        print(f'Failed to scrape {business_url}:', e)
        return {'name': 'N/A', 'ratings': 'N/A', 'numberOfReviews': 'N/A', 'phone': 'N/A', 'website': 'N/A', 'address': 'N/A', 'link': business_url}


def save_to_csv(data, filename):
    with open(filename, mode='w', newline='', encoding='utf-8') as file:
        writer = csv.DictWriter(file, fieldnames=['name', 'ratings', 'numberOfReviews', 'phone', 'website', 'address', 'link'])
        writer.writeheader()
        for item in data:
            writer.writerow(item)
    print(f'Data saved to {filename}')


def update_city_status(city_name, excel_filename):
    wb = openpyxl.load_workbook(excel_filename)
    sheet = wb.active
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == city_name:
            sheet.cell(row=row, column=2).value = 'Done'
            break
    wb.save(excel_filename)


def main():
    excel_filename = 'cities.xlsx'
    wb = openpyxl.load_workbook(excel_filename)
    sheet = wb.active
    business_types = [
        "Mechanics", "Cleaners", "Car detailing service", "Car Wash", "Electrician"
        "Plumbers", "construction", "Salons", "Roofing","Landscaping", "Retail Stores", 
        "Real Estate", "Warehouse", "Property Management","Architecture"
    ]
    # business_types = ['Warehouse','Property Management','Healthcare','Vets','Cooperate Offices']

    for row in range(2, sheet.max_row + 1):
        city = sheet.cell(row=row, column=1).value
        status = sheet.cell(row=row, column=2).value
        
        if status != 'Done':
            for business_type in business_types:
                search_query = f"{business_type} in {city} US"
                print(search_query)
                output_filename = f"{business_type.replace(' ', '')}in{city}.csv"
                try:
                    scrape_google_maps_data(search_query, output_filename)
                except Exception as e:
                    print(f"Error occurred for {business_type} in {city}: {e}")

            update_city_status(city, excel_filename)


if __name__ == "__main__":
    main()
